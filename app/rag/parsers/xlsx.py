"""XLSX parser — read-only, data-only cell text via openpyxl."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.exceptions.base import RAGDocumentValidationError
from app.rag.parsers.base import ParsedDocument, SourceFormat

# Fixed safety ceilings (parser layer): config may only tighten these.
_MAX_TEXT_CHARACTERS = 5_000_000
_MAX_ROWS_PER_SHEET = 65_536


class XlsxParser:
    """Extract non-empty cell text from all sheets.

    read_only + data_only keep memory bounded; rows are joined with tabs,
    sheets separated by blank lines.  Formulas are evaluated to cached
    values (data_only).  Content judgement belongs to the safety layer.
    """

    def parse(self, filename: str, content: bytes) -> ParsedDocument:
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise RAGDocumentValidationError(
                "无法解析 XLSX 文档：文件损坏或格式不受支持。"
            ) from exc
        try:
            sections: list[str] = []
            for worksheet in workbook.worksheets:
                rows: list[str] = []
                for index, row in enumerate(
                    worksheet.iter_rows(values_only=True), start=1
                ):
                    if index > _MAX_ROWS_PER_SHEET:
                        break
                    cells = [
                        str(cell).strip()
                        for cell in row
                        if cell is not None and str(cell).strip()
                    ]
                    if cells:
                        rows.append("\t".join(cells))
                if rows:
                    sections.append("\n".join(rows))
        finally:
            workbook.close()
        text = "\n\n".join(sections)
        if len(text) > _MAX_TEXT_CHARACTERS:
            raise RAGDocumentValidationError("XLSX 文档内容超过限制。")
        return ParsedDocument(
            filename=filename,
            text=text,
            source_format=_SOURCE_FORMAT,
            page_count=None,
        )


_SOURCE_FORMAT: SourceFormat = "xlsx"
